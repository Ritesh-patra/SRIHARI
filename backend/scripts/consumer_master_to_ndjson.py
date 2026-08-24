#!/usr/bin/env python3
"""Stream New Scope Consumer Master Excel (3 sheets) → NDJSON.

Uses openpyxl read_only. Does not full-count huge sheets up-front.
Normalizes sheet-specific column quirks into a common row shape.
"""

from __future__ import annotations

import json
import sys
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
IMPORTS = ROOT / "storage" / "app" / "imports"

DEFAULT_XLSX = IMPORTS / "New_Scope_Revised_Updated_Consumer_Master_List.xlsx"
OUT_PATH = IMPORTS / "consumers_scope_master.ndjson"

# Skip empty / zone-list helper sheets
SKIP_SHEETS = {"sheet1"}


def split_code_name(value) -> tuple[str | None, str | None]:
    if value is None:
        return None, None
    text = str(value).strip()
    if text == "":
        return None, None
    if "|" in text:
        left, right = text.split("|", 1)
        code = left.strip() or None
        name = right.strip() or None
        return code, name
    return text, None


def cell(row: dict, *keys: str) -> str | None:
    for k in keys:
        v = row.get(k)
        if v is None:
            continue
        s = str(v).strip()
        if s != "":
            return s
    return None


def normalize_row(raw: dict, sheet: str) -> dict | None:
    region = cell(raw, "Region", "region")
    circle = cell(raw, "circle_name", "Circle", "circle")
    division = cell(raw, "div_name", "Division", "division")
    zone = cell(raw, "zone_name", "Zone", "zone")

    # Urban sheet: dtr_name holds code; unnamed col after it holds name
    dtr_code, dtr_name = split_code_name(raw.get("dtr_name") or raw.get("DTR Code") or raw.get("DTR Name"))
    if dtr_name is None:
        dtr_name = cell(raw, "dtr_name_text", "DTR Name")
    if dtr_code is None:
        # sometimes only name-like values
        maybe = cell(raw, "dtr_name")
        if maybe and maybe.isdigit():
            dtr_code = maybe

    feeder_code = cell(raw, "Feeder Code", "feeder_code", "Feeder code")
    feeder_name = cell(raw, "Feeder name", "Feeder Name", "feeder_name")
    if feeder_code is None and feeder_name:
        # Non-contagious: feeder_name is CODE|NAME
        fc, fn = split_code_name(feeder_name)
        if fc and (fn is not None or "|" in str(raw.get("feeder_name") or "")):
            feeder_code, feeder_name = fc, fn or fc
        elif fc and fc.isdigit():
            feeder_code, feeder_name = fc, fn or fc

    # If feeder_name still looks like CODE|NAME
    if feeder_name and "|" in feeder_name and feeder_code is None:
        feeder_code, feeder_name = split_code_name(feeder_name)

    ivrs = cell(raw, "IVRS Number", "ivrs", "IVRS")
    msn = cell(raw, "Meter Serial Number", "meter_serial_no", "New Meter Serial Number", "msn")
    name = cell(raw, "consumer_name", "Consumer Name", "name")
    address = cell(raw, "address", "Address")
    phone = cell(raw, "Mobile Number", "phone", "mobile")
    phase = cell(raw, "phase", "Meter Type", "New Meter Type")

    if not ivrs and not msn and not name:
        return None
    if not dtr_code and not feeder_code:
        # Still keep if we have IVRS — orphan under placeholder later
        pass

    return {
        "sheet": sheet,
        "Region": region,
        "Circle": circle,
        "Division": division,
        "Zone": zone,
        "Substation Name": zone,  # no SS column in this workbook
        "Feeder Code": feeder_code,
        "Feeder Name": feeder_name or feeder_code,
        "DTR Code": dtr_code,
        "DTR Name": dtr_name or dtr_code,
        "IVRS Number": ivrs,
        "New Meter Serial Number": msn,
        "Consumer Name": name,
        "Mobile Number": phone,
        "address": address,
        "phase": phase,
    }


def iter_sheet(ws, sheet_name: str):
    rows = ws.iter_rows(values_only=True)
    try:
        header_row = next(rows)
    except StopIteration:
        return

    headers: list[str] = []
    empty_after_dtr = False
    for i, h in enumerate(header_row):
        label = "" if h is None else str(h).strip()
        if label == "" and i > 0:
            prev = headers[i - 1] if headers else ""
            # Urban sheet: blank header after dtr_name = DTR display name
            if prev.lower() in {"dtr_name", "dtr name", "dtr code"}:
                label = "dtr_name_text"
                empty_after_dtr = True
            else:
                label = f"col_{i}"
        headers.append(label)

    for row in rows:
        if row is None:
            continue
        raw: dict = {}
        any_val = False
        for i, h in enumerate(headers):
            val = row[i] if i < len(row) else None
            if val is not None and str(val).strip() != "":
                any_val = True
            raw[h] = val
        if not any_val:
            continue
        norm = normalize_row(raw, sheet_name)
        if norm:
            yield norm


def main() -> int:
    xlsx = Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT_XLSX
    out = Path(sys.argv[2]) if len(sys.argv) > 2 else OUT_PATH
    if not xlsx.is_file():
        print(f"Missing Excel: {xlsx}", file=sys.stderr)
        return 1

    IMPORTS.mkdir(parents=True, exist_ok=True)
    print(f"Reading {xlsx.name} …", flush=True)
    wb = load_workbook(xlsx, read_only=True, data_only=True)
    total = 0
    per_sheet: dict[str, int] = {}

    with out.open("w", encoding="utf-8") as fh:
        for sheet_name in wb.sheetnames:
            if sheet_name.strip().lower() in SKIP_SHEETS:
                print(f"  skip sheet: {sheet_name}", flush=True)
                continue
            ws = wb[sheet_name]
            n = 0
            print(f"  sheet: {sheet_name}", flush=True)
            for row in iter_sheet(ws, sheet_name):
                fh.write(json.dumps(row, ensure_ascii=False) + "\n")
                n += 1
                total += 1
                if n % 50000 == 0:
                    print(f"    … {n}", flush=True)
            per_sheet[sheet_name] = n
            print(f"    wrote {n}", flush=True)

    wb.close()
    print(f"Done -> {out} ({total} rows)", flush=True)
    for k, v in per_sheet.items():
        print(f"  {k}: {v}", flush=True)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
