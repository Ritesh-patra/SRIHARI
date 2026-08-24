#!/usr/bin/env python3
"""Stream Feeder / DT / Consumer master Excel → NDJSON for seas:import-masters.

Uses openpyxl read_only. Writes one JSON object per line (no giant JSON array).
Does not count rows up-front — streams and prints progress every 50k rows.
"""

from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
IMPORTS = ROOT / "storage" / "app" / "imports"

FEEDER_ALIASES = {
    "region": "Region",
    "circle": "Circle",
    "division": "Division",
    "zone": "Zone",
    "zone/dc": "Zone",
    "dc": "Zone",
    "substation name": "Substation Name",
    "substation": "Substation Name",
    "sub station name": "Substation Name",
    "sub-station name": "Substation Name",
    "ss code": "Substation Code",
    "substation code": "Substation Code",
    "sub-station code": "Substation Code",
    "feeder code": "Feeder Code",
    "feeder name": "Feeder Name",
    "feeder": "Feeder Name",
}

DTR_ALIASES = {
    **FEEDER_ALIASES,
    "dt location code": "DTR Code",
    "dtr code": "DTR Code",
    "location_name": "DTR Name",
    "location name": "DTR Name",
    "dtr name": "DTR Name",
    "dtr": "DTR Name",
    "capacity": "DTR Capacity in kVA",
    "dtr capacity in kva": "DTR Capacity in kVA",
}

CONSUMER_ALIASES = {
    **DTR_ALIASES,
    "consumer name": "Consumer Name",
    "mobile number": "Mobile Number",
    "ivrs number": "IVRS Number",
    "new meter serial number": "New Meter Serial Number",
    "new meter type": "New Meter Type",
    "msn": "New Meter Serial Number",
}


def norm_header(h) -> str:
    if h is None:
        return ""
    return str(h).strip()


def header_key(h: str) -> str:
    key = norm_header(h).lower().replace("_", " ").replace("-", " ")
    return " ".join(key.split())


def map_headers(raw_headers, aliases: dict[str, str]) -> list[str | None]:
    mapped: list[str | None] = []
    for i, h in enumerate(raw_headers):
        key = header_key(h) if h is not None else f"col_{i}"
        mapped.append(aliases.get(key))
    return mapped


def norm_cell(value):
    if value is None:
        return None
    if isinstance(value, bool):
        return value
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        return str(value)
    text = str(value).strip()
    return text if text != "" else None


def capacity_cell(value):
    if value is None or value == "":
        return None
    try:
        num = float(value)
        return int(num) if num.is_integer() else num
    except (TypeError, ValueError):
        return norm_cell(value)


def find_sheet(wb, preferred: list[str]):
    names = wb.sheetnames
    for want in preferred:
        if want in names:
            return want
        match = next((s for s in names if s.strip() == want.strip()), None)
        if match:
            return match
    return names[0] if names else None


def convert_sheet(
    xlsx: Path,
    out_path: Path,
    aliases: dict[str, str],
    sheet_prefs: list[str],
    label: str,
) -> int:
    print(f"[{label}] Opening {xlsx.name} ...", flush=True)
    wb = load_workbook(xlsx, read_only=True, data_only=True)
    sheet_name = find_sheet(wb, sheet_prefs)
    if sheet_name is None:
        print(f"[{label}] No sheets in {xlsx}", file=sys.stderr)
        wb.close()
        return 0

    ws = wb[sheet_name]
    print(f"[{label}] Sheet={sheet_name!r}", flush=True)
    rows = ws.iter_rows(values_only=True)
    raw_headers = next(rows, None)
    if not raw_headers:
        wb.close()
        return 0

    print(f"[{label}] Headers: {tuple(norm_header(h) for h in raw_headers)}", flush=True)
    mapped = map_headers(raw_headers, aliases)
    print(f"[{label}] Mapped: {[m for m in mapped if m]}", flush=True)

    out_path.parent.mkdir(parents=True, exist_ok=True)
    count = 0
    sample_printed = False

    with out_path.open("w", encoding="utf-8") as f:
        for row in rows:
            if row is None or all(c is None or (isinstance(c, str) and not c.strip()) for c in row):
                continue
            item: dict = {}
            for i, canon in enumerate(mapped):
                if canon is None:
                    continue
                val = row[i] if i < len(row) else None
                if canon == "DTR Capacity in kVA":
                    cell = capacity_cell(val)
                else:
                    cell = norm_cell(val)
                if canon in item and item[canon] is not None:
                    continue
                item[canon] = cell

            if not item.get("Region"):
                continue
            if label == "feeders" and not item.get("Feeder Code"):
                continue
            if label == "dtrs" and not item.get("DTR Code"):
                continue
            if label == "consumers" and not (
                item.get("DTR Code") or item.get("IVRS Number") or item.get("Feeder Code")
            ):
                continue

            if not sample_printed:
                print(f"[{label}] Sample: {item}", flush=True)
                sample_printed = True

            f.write(json.dumps(item, ensure_ascii=False, default=str))
            f.write("\n")
            count += 1
            if count % 50000 == 0:
                print(f"[{label}] ... {count} rows", flush=True)

    wb.close()
    print(f"[{label}] Wrote {count} rows -> {out_path.name}", flush=True)
    return count


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--imports", type=Path, default=IMPORTS)
    parser.add_argument("--feeder-xlsx", type=Path, default=None)
    parser.add_argument("--dtr-xlsx", type=Path, default=None)
    parser.add_argument("--consumer-xlsx", type=Path, default=None)
    args = parser.parse_args()
    imports: Path = args.imports
    imports.mkdir(parents=True, exist_ok=True)

    feeder_xlsx = args.feeder_xlsx or imports / "Feeder_Master_Final.xlsx"
    dtr_xlsx = args.dtr_xlsx or imports / "DT_Master_Final.xlsx"
    consumer_xlsx = args.consumer_xlsx or imports / "Consumer_MI_Done_Till_28_July_2026.xlsx"

    for path, name in [
        (feeder_xlsx, "feeder"),
        (dtr_xlsx, "dtr"),
        (consumer_xlsx, "consumer"),
    ]:
        if not path.exists():
            print(f"Missing {name} Excel: {path}", file=sys.stderr)
            return 1

    n1 = convert_sheet(
        feeder_xlsx,
        imports / "feeders_master.ndjson",
        FEEDER_ALIASES,
        ["Sheet1"],
        "feeders",
    )
    n2 = convert_sheet(
        dtr_xlsx,
        imports / "dtrs_master.ndjson",
        DTR_ALIASES,
        ["Sheet1"],
        "dtrs",
    )
    n3 = convert_sheet(
        consumer_xlsx,
        imports / "consumers_master.ndjson",
        CONSUMER_ALIASES,
        ["Consumer MI ", "Consumer MI"],
        "consumers",
    )

    print(f"TOTAL feeders={n1} dtrs={n2} consumers={n3}", flush=True)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
