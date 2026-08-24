#!/usr/bin/env python3
"""Convert Sample_Data.xlsx sheets into JSON files for seas:import-sample-data."""

from __future__ import annotations

import json
import sys
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
IMPORTS = ROOT / "storage" / "app" / "imports"
XLSX = IMPORTS / "Sample_Data.xlsx"

SHEETS = {
    "Feeder MI ": "feeders.json",
    "DTR MI ": "dtrs.json",
    "Consumer MI ": "consumers.json",
}


def norm_cell(value):
    if value is None:
        return None
    if isinstance(value, bool):
        return value
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        return str(value)
    text = str(value).strip()
    return text if text != "" else None


def sheet_to_rows(ws) -> list[dict]:
    rows_iter = ws.iter_rows(values_only=True)
    raw_headers = next(rows_iter, None)
    if not raw_headers:
        return []
    headers = [str(h).strip() if h is not None else f"col_{i}" for i, h in enumerate(raw_headers)]
    out: list[dict] = []
    for row in rows_iter:
        if row is None or all(c is None or (isinstance(c, str) and not c.strip()) for c in row):
            continue
        item = {}
        for i, header in enumerate(headers):
            val = row[i] if i < len(row) else None
            # Keep capacity as number when possible; codes/names as strings.
            if header in {"DTR Capacity in kVA", "New MF", "New Meter MF", "Location Accuracy"}:
                if val is None or val == "":
                    item[header] = None
                else:
                    try:
                        num = float(val)
                        item[header] = int(num) if num.is_integer() else num
                    except (TypeError, ValueError):
                        item[header] = norm_cell(val)
            else:
                item[header] = norm_cell(val)
        # Skip rows without region (noise)
        if not item.get("Region"):
            continue
        out.append(item)
    return out


def main() -> int:
    if not XLSX.exists():
        print(f"Missing Excel file: {XLSX}", file=sys.stderr)
        return 1

    IMPORTS.mkdir(parents=True, exist_ok=True)
    wb = load_workbook(XLSX, read_only=True, data_only=True)

    for sheet_name, out_name in SHEETS.items():
        if sheet_name not in wb.sheetnames:
            # try trimmed match
            match = next((s for s in wb.sheetnames if s.strip() == sheet_name.strip()), None)
            if match is None:
                print(f"Sheet not found: {sheet_name!r}. Available: {wb.sheetnames!r}", file=sys.stderr)
                wb.close()
                return 1
            sheet_name = match

        rows = sheet_to_rows(wb[sheet_name])
        out_path = IMPORTS / out_name
        with out_path.open("w", encoding="utf-8") as f:
            json.dump(rows, f, ensure_ascii=False, indent=None)
        print(f"Wrote {len(rows)} rows -> {out_path}")

    wb.close()
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
